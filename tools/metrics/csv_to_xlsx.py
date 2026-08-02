#!/usr/bin/env python3
"""Consolida os CSVs de tools/metrics/run_extract.sh num único .xlsx
(uma aba por query) — mais prático pra abrir num add-in do Excel do que
5 arquivos CSV soltos.

Uso:
    python3 tools/metrics/csv_to_xlsx.py reports/metrics/2026-08
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def main() -> int:
    if len(sys.argv) != 2:
        print("Uso: python3 tools/metrics/csv_to_xlsx.py <diretório com os q*.csv>", file=sys.stderr)
        return 1

    src_dir = Path(sys.argv[1])
    csv_files = sorted(src_dir.glob("q*.csv"))
    if not csv_files:
        print(f"Nenhum q*.csv encontrado em {src_dir}", file=sys.stderr)
        return 1

    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)  # remove a aba default vazia

    for csv_path in csv_files:
        sheet_name = csv_path.stem[:31]  # limite de 31 caracteres do Excel
        ws = wb.create_sheet(sheet_name)
        with csv_path.open(encoding="utf-8", newline="") as f:
            for row in csv.reader(f):
                ws.append(row)
        # Autofit aproximado — largura pela maior célula de cada coluna.
        for col_idx in range(1, ws.max_column + 1):
            width = max((len(str(cell.value or "")) for cell in ws[get_column_letter(col_idx)]), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 40)

    out_path = src_dir / "cockpit_unit_economics.xlsx"
    wb.save(out_path)
    print(f"Gerado: {out_path} ({len(csv_files)} abas)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
