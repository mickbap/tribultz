"""Contrato do catálogo CFOP versionado (#688).

Três eixos, na ordem que a issue exige: identidade do artefato → semântica da
propriedade → encadeamento (mudar o artefato tem de ser observável na
propriedade). O terceiro elo da cadeia — a regra I08-191 — é a #615 e ainda não
existe; o que este arquivo garante é que o elo anterior é observável.
"""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import pathlib

import pytest
from openpyxl import load_workbook

from app.data import cfop_table as t
from app.data.provenance import ArtifactProvenance

DATA = pathlib.Path(t.__file__).parent


class TestIdentidadeDoArtefato:
    def test_proveniencia_tem_o_contrato_completo(self):
        p = t.provenance()
        assert isinstance(p, ArtifactProvenance)
        for campo in ("artefato", "versao", "fonte", "source_url", "observado_em", "fingerprint"):
            assert campo in p.to_dict()

    def test_fonte_e_o_portal_oficial(self):
        # ArtifactProvenance já recusaria host não oficial; aqui firmamos QUAL portal.
        assert "nfe.fazenda.gov.br" in t.provenance().source_url
        assert "nfe.fazenda.gov.br" in t.instituido_por().source_url

    def test_fingerprint_bate_com_o_arquivo_bruto_preservado(self):
        """Auditoria byte a byte: o hash registrado tem de corresponder ao XLSX
        oficial guardado no repo. É isso que impede uma edição do JSON de passar
        por 'dado oficial'."""
        meta = json.loads((DATA / "cfop_table.json").read_text(encoding="utf-8"))["meta"]
        bruto = DATA / meta["arquivo_bruto"]
        assert bruto.exists(), "o artefato bruto sumiu — a proveniência vira alegação"
        assert hashlib.sha256(bruto.read_bytes()).hexdigest() == t.provenance().fingerprint

    def test_tabela_e_fonte_sem_versao_propria_mas_o_IT_tem(self):
        # A Tabela de CFOP é publicada por data, não por versão. Quem carrega
        # versão é o Informe Técnico que instituiu a coluna.
        assert t.provenance().is_live_source is True
        assert t.instituido_por().versao == "2.10"
        assert t.instituido_por().artefato == "IT 2023.002"

    def test_it_v210_bruto_tem_fingerprint_auditavel(self):
        meta = json.loads((DATA / "cfop_table.json").read_text(encoding="utf-8"))["meta"]
        it = meta["instituido_por"]
        assert hashlib.sha256((DATA / it["arquivo_bruto"]).read_bytes()).hexdigest() == it["fingerprint"]

    def test_v200_permanece_no_historico_com_artefato_e_vigencia(self):
        anterior, = t.historico()
        assert anterior["instituido_por"]["versao"] == "2.00"
        assert (DATA / anterior["arquivo_bruto"]).exists()
        assert anterior["vigencia_indExcIBSCBS"] == {
            "homologacao": "2026-09-01",
            "producao": "2026-11-03",
        }

    def test_v210_nao_inventa_data_de_aplicacao(self):
        assert t.aplicacao_v210() == {
            "homologacao": None,
            "producao": None,
            "texto_oficial": "Não aplicável",
            "efeito": (
                "Inclusão exclusivamente informativa de Título e Descrição; "
                "sem nova regra ou rejeição."
            ),
        }


class TestContratoDeContagem:
    def test_619_547_72(self):
        """Contrato fechado pelo parecer: 619 totais, 547 com 0, 72 com 1."""
        c = t.contagem()
        assert c == {"total": 619, "indExcIBSCBS_0": 547, "indExcIBSCBS_1": 72}
        assert c["indExcIBSCBS_0"] + c["indExcIBSCBS_1"] == c["total"]

    def test_contagem_declarada_bate_com_o_dado_carregado(self):
        # meta não pode divergir do conteúdo: seria proveniência mentindo
        c = t.contagem()
        assert len(t.all_cfops()) == c["total"]
        assert len(t.cfops_permitidos_contribuinte_exclusivo()) == c["indExcIBSCBS_1"]


class TestDiffV200V210:
    @staticmethod
    def _operacional(arquivo: pathlib.Path) -> dict[str, tuple]:
        planilha = load_workbook(arquivo, data_only=True, read_only=True)["CFOP"]
        return {
            str(row[0]).strip().zfill(4): tuple(
                value.date() if isinstance(value, dt.datetime) else value
                for value in row[:12]
            )
            for row in planilha.iter_rows(min_row=2, max_col=14, values_only=True)
            if row[0] is not None
        }

    def test_619_codigos_e_12_colunas_operacionais_sao_identicos(self):
        anterior, = t.historico()
        atual = json.loads((DATA / "cfop_table.json").read_text(encoding="utf-8"))["meta"]
        v200 = self._operacional(DATA / anterior["arquivo_bruto"])
        v210 = self._operacional(DATA / atual["arquivo_bruto"])

        assert len(v200) == 619
        assert v210 == v200

    def test_v210_acrescenta_exatamente_as_duas_colunas_explicativas(self):
        atual = json.loads((DATA / "cfop_table.json").read_text(encoding="utf-8"))["meta"]
        planilha = load_workbook(
            DATA / atual["arquivo_bruto"], data_only=True, read_only=True
        )["CFOP"]
        cabecalho = [
            cell.value
            for cell in next(planilha.iter_rows(min_row=1, max_row=1, max_col=14))
        ]
        assert cabecalho[12:] == ["Título do CFOP", "Descrição (nota explicativa)"]


class TestConflito84x72:
    def test_conflito_permanece_UNRESOLVED(self):
        k = t.conflito_contagem()
        assert k["conflict_status"] == "UNRESOLVED"

    def test_ambas_as_contagens_ficam_registradas(self):
        k = t.conflito_contagem()
        assert k["it_2023_002_v200"]["textual_count"] == 84
        assert k["tabela_oficial_20260825"]["observed_count"] == 72

    def test_nao_geramos_os_12_codigos_que_fechariam_84(self):
        """A tentação óbvia — inventar 12 CFOPs para bater com o texto do IT."""
        assert len(t.cfops_permitidos_contribuinte_exclusivo()) == 72

    def test_o_lookup_usa_o_valor_individual_publicado_e_nao_a_contagem(self):
        """Cada CFOP responde pelo próprio indicador. Contagem agregada é
        estatística; ela não decide nada, e o conflito nela não contamina o
        lookup de um código específico."""
        for c in sorted(t.all_cfops())[:40]:
            reg = t.get(c)
            assert reg is not None
            assert t.ind_exc_ibscbs(c) == reg["indExcIBSCBS"]


class TestDominioCompleto:
    def test_representa_a_tabela_inteira_e_nao_o_subconjunto_permitido(self):
        """O defeito que a #688 corrige é justamente guardar só os permitidos."""
        assert len(t.all_cfops()) == 619
        assert len(t.cfops_permitidos_contribuinte_exclusivo()) == 72
        assert len(t.all_cfops()) > len(t.cfops_permitidos_contribuinte_exclusivo())

    def test_todo_codigo_tem_4_digitos(self):
        assert all(len(c) == 4 and c.isdigit() for c in t.all_cfops())

    def test_todas_as_propriedades_operacionais_e_explicativas_por_codigo(self):
        esperado = {"inicio_vigencia", "fim_vigencia", "indNFe", "indComunica", "indTransp",
                    "indDevol", "indRetor", "indAnula", "indRemes", "indComb", "indExcIBSCBS",
                    "titulo", "descricao"}
        reg = t.get("5102")
        assert reg is not None, "5102 tem de existir no domínio oficial"
        assert set(reg) == esperado

    def test_titulo_e_descricao_sao_metadados_presentes_nos_619_cfops(self):
        for codigo in t.all_cfops():
            reg = t.get(codigo)
            assert reg is not None
            assert reg["titulo"].strip()
            assert reg["descricao"].strip()


class TestSemanticaDoIndicador:
    def test_valores_sao_apenas_0_ou_1(self):
        assert {t.ind_exc_ibscbs(c) for c in t.all_cfops()} == {"0", "1"}

    def test_cfop_desconhecido_e_None_nao_False(self):
        """`None` (fora do domínio oficial) é resposta DIFERENTE de `False`
        (existe e não é permitido). Colapsar as duas apagaria a distinção entre
        'não sei' e 'sei que não'."""
        assert t.permitido_contribuinte_exclusivo_ibscbs("9999") is None
        assert t.ind_exc_ibscbs("9999") is None

    def test_permitido_e_nao_permitido_sao_ambos_conhecidos(self):
        permitidos = t.cfops_permitidos_contribuinte_exclusivo()
        um_permitido = sorted(permitidos)[0]
        um_negado = sorted(t.all_cfops() - permitidos)[0]
        assert t.permitido_contribuinte_exclusivo_ibscbs(um_permitido) is True
        assert t.permitido_contribuinte_exclusivo_ibscbs(um_negado) is False

    def test_nenhuma_api_sugere_inferencia_proibida(self):
        """O indicador responde admissibilidade. Não pode existir função que
        convide a derivar incidência, isenção, crédito ou conformidade."""
        proibidos = ("incidencia", "isencao", "imunidade", "credito", "debito",
                     "tributad", "compliance", "conformidade")
        publicos = [n for n in dir(t) if not n.startswith("_")]
        for nome in publicos:
            assert not any(p in nome.lower() for p in proibidos), (
                f"API '{nome}' sugere inferência proibida a partir de indExcIBSCBS"
            )


class TestLimiteTemporal:
    @pytest.mark.parametrize("quando,esperado", [
        (dt.date(2026, 8, 31), False),   # antes da homologação
        (dt.date(2026, 9, 1), False),    # homologação: ainda informativo
        (dt.date(2026, 11, 2), False),   # véspera da produção
        (dt.date(2026, 11, 3), True),    # produção
        (dt.date(2027, 1, 1), True),
    ])
    def test_efeito_de_rejeicao_so_a_partir_da_producao(self, quando, esperado):
        """O próprio artefato declara caráter informativo até a produção.
        Rejeitar antes seria endurecer além do que a fonte autoriza."""
        assert t.efeito_de_rejeicao_em(quando) is esperado


class TestEncadeamento:
    def test_mudar_o_artefato_e_observavel_na_propriedade(self, tmp_path, monkeypatch):
        """Elo artefato → propriedade. Sem isto o catálogo é hardcode com mais
        cerimônia (critério de aceite da #688)."""
        doc = json.loads((DATA / "cfop_table.json").read_text(encoding="utf-8"))
        alvo = sorted(t.cfops_permitidos_contribuinte_exclusivo())[0]
        assert t.permitido_contribuinte_exclusivo_ibscbs(alvo) is True

        doc["cfop"][alvo]["indExcIBSCBS"] = "0"
        doc["meta"]["fingerprint"] = "0" * 64
        falso = tmp_path / "cfop_table.json"
        falso.write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")

        monkeypatch.setattr(t, "_ARQUIVO", falso)
        t._doc.cache_clear()
        try:
            assert t.permitido_contribuinte_exclusivo_ibscbs(alvo) is False
            assert t.provenance().fingerprint == "0" * 64
        finally:
            monkeypatch.undo()   # o arquivo real volta ANTES da reverificação
            t._doc.cache_clear()

        # e o catálogo real permanece intacto
        assert t.permitido_contribuinte_exclusivo_ibscbs(alvo) is True
        assert t.provenance().fingerprint != "0" * 64


class TestNaoConfundirComOutraRegra:
    def test_a_allowlist_do_DANFE_T2_nao_e_este_catalogo(self):
        """`DANFE_SIMPLIFICADO_CFOP` (I08-150/Rejeição 725, NT 2026.002) usa 10
        CFOPs de venda direta ao consumidor. Assunto DIFERENTE de indExcIBSCBS
        (IT 2023.002 / NT 2026.007). Conflar os dois inventaria regra."""
        danfe_t2 = {"5101", "5102", "5103", "5104", "5115", "5405", "5656", "5667", "5910", "5933"}
        assert danfe_t2 != t.cfops_permitidos_contribuinte_exclusivo()
        assert danfe_t2 <= t.all_cfops(), "os 10 do DANFE T2 existem no domínio oficial"


class TestConflitoNaoCondicionaComportamento:
    """O conflito 84×72 é registro documental, não gate.

    Round Fiscal 27/08-D canonizou a Tabela de 25/08 como domínio operacional.
    O conflito segue UNRESOLVED — a causa documental continua aberta — mas não
    bloqueia lookup nem determinismo.
    """

    def test_efeito_operacional_declara_que_nao_bloqueia(self):
        e = t.conflito_contagem()["efeito_operacional"]
        assert e["bloqueia_lookup"] is False
        assert e["bloqueia_determinismo_i08_191"] is False
        assert e["dominio_operacional"] == 72

    def test_unresolved_preservado(self):
        assert t.conflito_contagem()["conflict_status"] == "UNRESOLVED"

    def test_lookup_responde_com_conflito_aberto(self):
        assert t.conflito_contagem()["conflict_status"] == "UNRESOLVED"
        assert t.permitido_contribuinte_exclusivo_ibscbs(
            sorted(t.cfops_permitidos_contribuinte_exclusivo())[0]) is True

    def test_nenhum_requisito_conflict_resolved_no_codigo(self):
        import pathlib
        for mod in ("cfop_table.py", "../services/rules_i08_191.py"):
            src = (pathlib.Path(t.__file__).parent / mod).read_text(encoding="utf-8")
            assert "conflict_84_72_resolved" not in src
            assert "conflito_resolvido" not in src
